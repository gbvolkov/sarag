#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Builds a SQLite table from the Excel sheet using EXACTLY 26 predefined
SQL-friendly column names. Date columns are normalized to ISO 8601 (YYYY-MM-DD)
and stored with DATE affinity (TEXT). Robust header/column detection scans ALL
columns and keeps any column that has data (prevents dropping the leftmost ID).
"""
import sqlite3
import pandas as pd
import numpy as np
import re
from openpyxl import load_workbook
from datetime import datetime, date, timedelta
from typing import List, Tuple

# ========= CONFIG =========
XLSX_PATH  = "./data/tasks.xlsx"  # <- set path
DB_PATH    = "./data/tasks.db"
TABLE_NAME = "tasks"
# ==========================

RAW_TO_SQL = {
    "Номер задачи в Битрикс": "bitrix_task_id",
    "Номер ЗНИ": "request_id",
    "Текущий приоритет обращения": "current_ticket_priority",
    "Группа приоритетов": "priority_group",
    "Наименование": "title",
    "Проект Битрикс": "bitrix_project",
    "Инициатор": "initiator",
    "Куратор": "curator",
    "Текущий исполнитель": "current_assignee",
    "ЗНИ текущий этап": "request_current_stage",
    "Статус ЗНИ": "request_status",
    "Дата заведения в системе": "created_date",
    "Плановая дата окончания бизнес-анализа": "planned_business_analysis_end_date",
    "Плановая дата окончания анализа": "planned_analysis_end_date",
    "Плановая дата выполнения": "planned_completion_date",
    "Дата завершения": "completed_date",
    "Код завершения ЗНИ": "request_completion_code",
    "SLA — Дата начала SLA": "sla_start_date",
    "SLA — SLA (кал. дней)": "sla_calendar_days",
    "SLA — Норматив SLA (кал. дней)": "sla_calendar_days_target",
    "SLA — SLA (раб. дней)": "sla_work_days",
    "SLA — Норматив SLA (раб. дней)": "sla_work_days_target",
    "Дата оценки": "estimate_date",
    "Общая оценка, час": "estimate_total_hours",
    "Уточненная оценка, час": "estimate_refined_hours",
    "Факт, час": "actual_hours",
}

EN_COLUMNS_ORDER = [
    "bitrix_task_id","request_id","current_ticket_priority","priority_group","title",
    "bitrix_project","initiator","curator","current_assignee","request_current_stage",
    "request_status","created_date","planned_business_analysis_end_date","planned_analysis_end_date",
    "planned_completion_date","completed_date","request_completion_code","sla_start_date",
    "sla_calendar_days","sla_calendar_days_target","sla_work_days","sla_work_days_target",
    "estimate_date","estimate_total_hours","estimate_refined_hours","actual_hours",
]

DATE_COLS = [
    "created_date","planned_business_analysis_end_date","planned_analysis_end_date",
    "planned_completion_date","completed_date","sla_start_date","estimate_date"
]

def cell(ws, r, c):
    v = ws.cell(r, c).value
    if isinstance(v, str):
        v = v.strip()
        if v == "":
            return None
    return v

def detect_headers_and_columns(ws):
    max_row, max_col = ws.max_row, ws.max_column
    # Pick header row by most text among rows 6..10 (robust to minor shifts)
    candidates = []
    for hr in range(5, 12):
        vals = [cell(ws, hr, c) for c in range(1, max_col+1)]
        nonnull = [x for x in vals if x is not None]
        if not nonnull: 
            continue
        txt = sum(isinstance(x, str) for x in nonnull)
        candidates.append((hr, txt))
    candidates.sort(key=lambda x: x[1], reverse=True)
    header_row = candidates[0][0] if candidates else 7
    subheader_row = header_row + 1
    data_start = subheader_row + 1

    # Build joint headers with carry-forward across ALL columns; keep columns that have any data
    carry = ""
    selected = []  # (col_idx, joint_header)
    for c in range(1, max_col+1):
        p = cell(ws, header_row, c)
        s = cell(ws, subheader_row, c)
        if isinstance(p, str) and p.strip():
            carry = p.strip()
        if s and isinstance(s, str) and s.strip() and (not (isinstance(p, str) and s.strip().lower() == p.strip().lower())):
            joint = f"{carry} — {s.strip()}"
        else:
            joint = carry if (isinstance(p, str) and p.strip()) else (carry or (p if p is not None else None))
        has_data = any(
            cell(ws, r, c) is not None for r in range(data_start, max_row + 1)
        )
        if joint and has_data:
            selected.append((c, joint))
    return header_row, subheader_row, data_start, selected

def parse_date_iso(x):
    EXCEL_BASE = datetime(1899, 12, 30)
    if x is None or (isinstance(x, float) and np.isnan(x)): return None
    if isinstance(x, (datetime, pd.Timestamp)): return (x.date() if isinstance(x, datetime) else x.date()).isoformat()
    if isinstance(x, date): return x.isoformat()
    if isinstance(x, (int, float, np.integer, np.floating)):
        val = float(x)
        if 20000 <= val <= 80000:
            try:
                dt = EXCEL_BASE + timedelta(days=val)
                return dt.date().isoformat()
            except: return None
        if 946684800 <= val <= 2000000000:
            return datetime.utcfromtimestamp(val).date().isoformat()
        if 946684800000 <= val <= 2000000000000:
            return datetime.utcfromtimestamp(val/1000.0).date().isoformat()
        return None
    if isinstance(x, str):
        s = x.strip()
        if s == "": return None
        s_norm = re.sub(r"[\\/]", "-", s)
        for fmt in ("%Y-%m-%d","%Y.%m.%d","%d.%m.%Y","%d-%m-%Y","%d/%m/%Y","%m/%d/%Y","%m-%d-%Y"):
            try:
                return datetime.strptime(s_norm, fmt).date().isoformat()
            except: continue
        try:
            dt = pd.to_datetime(s, errors="coerce")
            if pd.notna(dt):
                return dt.date().isoformat()
        except: return None
    return None

def to_num(x):
    if x is None or (isinstance(x, float) and np.isnan(x)) or (isinstance(x, str) and x.strip()==""): return None
    if isinstance(x, (int,float,np.integer,np.floating)): return float(x)
    if isinstance(x, str):
        xx = x.replace(",", ".").replace(" ", "")
        try: return float(xx)
        except: return None
    return None

def main():
    wb = load_workbook(XLSX_PATH, data_only=True, read_only=False)
    ws = wb.worksheets[0]
    header_row, subheader_row, data_start, selected = detect_headers_and_columns(ws)
    raw_headers = [h for (_, h) in selected]

    # Validate mapping coverage
    missing = sorted(set(raw_headers) - set(RAW_TO_SQL.keys()))
    if missing:
        raise RuntimeError(f"Missing RU->EN mappings for headers: {missing}")

    # Extract data
    rows = []
    max_row, _ = ws.max_row, ws.max_column
    for r in range(data_start, max_row+1):
        vals = [cell(ws, r, c) for (c, _) in selected]
        if any(v is not None for v in vals):
            rows.append(vals)
    df_present = pd.DataFrame(rows, columns=[RAW_TO_SQL[h] for (_, h) in selected])
    df = df_present.reindex(columns=EN_COLUMNS_ORDER)

    # Convert types
    for col in DATE_COLS:
        if col in df.columns:
            df[col] = df[col].map(parse_date_iso)
    NUM_HINTS = ("hours","days","estimate","target","sla")
    for col in df.columns:
        if col not in DATE_COLS and any(h in col for h in NUM_HINTS):
            df[col] = df[col].map(to_num)

    # SQLite schema
    def sqlite_type(col: str, s: pd.Series) -> str:
        if col in DATE_COLS: return "DATE"
        s2 = s.dropna()
        if len(s2)==0: return "TEXT"
        if pd.api.types.is_integer_dtype(s2): return "INTEGER"
        return "REAL" if pd.api.types.is_float_dtype(s2) else "TEXT"

    con = sqlite3.connect(DB_PATH)
    cur = con.cursor()
    cur.execute(f'DROP TABLE IF EXISTS "{TABLE_NAME}"')
    col_defs = ", ".join([f'"{c}" {sqlite_type(c, df[c])}' for c in EN_COLUMNS_ORDER])
    cur.execute(f'CREATE TABLE "{TABLE_NAME}" ({col_defs})')
    placeholders = ", ".join(["?"]*len(EN_COLUMNS_ORDER))
    cur.executemany(f'INSERT INTO "{TABLE_NAME}" VALUES ({placeholders})', df.where(pd.notnull(df), None).values.tolist())
    con.commit()
    con.close()

    print(f"Created {DB_PATH} with table {TABLE_NAME}: {df.shape[0]} rows, {df.shape[1]} columns.")
    # Sanity check for bitrix_task_id
    non_null = int(df["bitrix_task_id"].notna().sum())
    print(f"bitrix_task_id non-null rows: {non_null}")

if __name__ == "__main__":
    main()
